from asyncore import write
from genericpath import exists
from heapq import merge
import pandas as pd
from io import BytesIO
from django.core.files.base import ContentFile
from django.db.models import Count

# from utility.custom_function import gabung_dictionary
# from unit_penjamin_mutu.custom_function import value_in_list_of_dict, multiple_value_exist,get_column, get_columns_from_worksheet,apply_style_to_cell, merge_cell
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
import time

# from .kelas import Nilai_raport

import os

current_folder = os.path.dirname(os.path.abspath(__file__))


def data_buku_induk(
    self,
    obj1,
    obj2,
    obj3,
    obj4,
    obj5,
    obj6,
    obj7,
    obj8,
    obj9,
    obj10,
    obj11,
    obj12,
    obj13,
):
    beasiswa = obj1
    nilai_x_1 = obj2
    ekskul_x_1 = obj3
    nilai_x_2 = obj4
    ekskul_x_2 = obj5
    nilai_xi_1 = obj6
    ekskul_xi_1 = obj7
    nilai_xi_2 = obj8
    ekskul_xi_2 = obj9
    nilai_xii_1 = obj10
    ekskul_xii_1 = obj11
    nilai_xii_2 = obj12
    ekskul_xii_2 = obj13

    wb_obj = openpyxl.load_workbook(os.path.join(current_folder, "template_induk.xlsx"))
    wb_obj.active = wb_obj["Data_Siswa"]
    sheet_obj = wb_obj.active
    nama_cell = sheet_obj.cell(row=2, column=4)
    nama_cell.value = str(self.NIS.NAMA)
    panggilan_cell = sheet_obj.cell(row=3, column=4)
    panggilan_cell.value = str(self.NAMA_PANGGILAN)
    jk_cell = sheet_obj.cell(row=4, column=4)
    jk_cell.value = str(self.NIS.JENIS_KELAMIN)
    ttl_cell = sheet_obj.cell(row=5, column=4)
    ttl_cell.value = str(
        str(self.NIS.TEMPAT_LAHIR) + ", " + str(self.NIS.TANGGAL_LAHIR)
    )
    agama_cell = sheet_obj.cell(row=6, column=4)
    agama_cell.value = str(self.NIS.AGAMA)
    kewarganegaraan_cell = sheet_obj.cell(row=7, column=4)
    kewarganegaraan_cell.value = str(self.KEWARGANEGARAAN)
    anakke_cell = sheet_obj.cell(row=8, column=4)
    anakke_cell.value = str(self.NIS.ANAK_KE)
    jumlah_saudara_kandung_cell = sheet_obj.cell(row=9, column=4)
    jumlah_saudara_kandung_cell.value = str(self.NIS.JUMLAH_SAUDARA_KANDUNG)
    jumlah_saudara_tiri_cell = sheet_obj.cell(row=10, column=4)
    jumlah_saudara_tiri_cell.value = str(self.JUMLAH_SAUDARA_TIRI)
    jumlah_saudara_angkat_cell = sheet_obj.cell(row=11, column=4)
    jumlah_saudara_angkat_cell.value = str(self.JUMLAH_SAUDARA_ANGKAT)
    yatim_cell = sheet_obj.cell(row=12, column=4)
    yatim_cell.value = str(self.ANAK_YATIM_PIATU)
    bahasa_cell = sheet_obj.cell(row=13, column=4)
    bahasa_cell.value = str(self.BAHASA)
    alamat_cell = sheet_obj.cell(row=16, column=4)
    alamat_cell.value = str(
        str(self.NIS.ALAMAT)
        + ", RT = "
        + str(self.NIS.RT)
        + " RW = "
        + str(self.NIS.RW)
        + ", "
        + str(self.NIS.DUSUN)
        + ", "
        + str(self.NIS.KELURAHAN)
        + ", "
        + str(self.NIS.KECAMATAN)
        + ". KODE POS ("
        + str(self.NIS.KODE_POS)
        + ")"
    )
    telp_cell = sheet_obj.cell(row=17, column=4)
    telp_cell.value = str(str(self.NIS.TELEPON) + "/" + str(self.NIS.HP))
    jenis_tinggal_cell = sheet_obj.cell(row=18, column=4)
    jenis_tinggal_cell.value = str(self.NIS.JENIS_TINGGAL)
    jarak_cell = sheet_obj.cell(row=19, column=4)
    jarak_cell.value = str(self.NIS.JARAK_RUMAH_KESEKOLAH_KM)
    golongan_darah_cell = sheet_obj.cell(row=22, column=4)
    golongan_darah_cell.value = str(self.GOLONGAN_DARAH)
    penyakit_cell = sheet_obj.cell(row=23, column=4)
    penyakit_cell.value = str(self.PENYAKIT_PERNAH_DIDERITA)
    kelainan_cell = sheet_obj.cell(row=24, column=4)
    kelainan_cell.value = str(self.KELAINAN_JASMANI)
    tinggi_berat_cell = sheet_obj.cell(row=25, column=4)
    tinggi_berat_cell.value = str(
        str(self.NIS.TINGGI_BADAN) + "/" + str(self.NIS.BERAT_BADAN)
    )
    tamatan_dari_cell = sheet_obj.cell(row=29, column=4)
    tamatan_dari_cell.value = str(self.TAMATAN_DARI)
    tgl_ijazah_cell = sheet_obj.cell(row=30, column=4)
    tgl_ijazah_cell.value = str(
        str(self.TANGGAL_IJAZAH_S) + " / " + str(self.NO_IJAZAH_S)
    )
    tgl_skhun_cell = sheet_obj.cell(row=31, column=4)
    tgl_skhun_cell.value = str(str(self.TANGGAL_SKHUN_S) + " / " + str(self.NO_SKHUN_S))
    lama_belajar_cell = sheet_obj.cell(row=32, column=4)
    lama_belajar_cell.value = str(self.LAMA_BELAJAR)
    pindahan_cell = sheet_obj.cell(row=34, column=4)
    pindahan_cell.value = str(self.PINDAHAN_DARI)
    alasan_pindah_cell = sheet_obj.cell(row=35, column=4)
    alasan_pindah_cell.value = str(self.ALASAN_PINDAHAN)
    diterima_dikelas_cell = sheet_obj.cell(row=37, column=4)
    diterima_dikelas_cell.value = str(self.DITERIMA_DI_KELAS)
    kelompok_cell = sheet_obj.cell(row=38, column=4)
    kelompok_cell.value = str(self.KELOMPOK)
    tanggal_diterima_cell = sheet_obj.cell(row=39, column=4)
    tanggal_diterima_cell.value = str(self.TANGGAL_DITERIMA)
    nama_ayah_cell = sheet_obj.cell(row=42, column=4)
    nama_ayah_cell.value = str(self.ORANG_TUA.NAMA_AYAH)
    ttl_ayah_cell = sheet_obj.cell(row=43, column=4)
    ttl_ayah_cell.value = str(
        str(self.ORANG_TUA.TEMPAT_LAHIR_AYAH)
        + ", "
        + str(self.ORANG_TUA.TAHUN_LAHIR_AYAH)
    )
    agama_ayah_cell = sheet_obj.cell(row=44, column=4)
    agama_ayah_cell.value = str(self.ORANG_TUA.AGAMA_AYAH)
    kewarganegaraan_ayah_cell = sheet_obj.cell(row=45, column=4)
    kewarganegaraan_ayah_cell.value = str(self.ORANG_TUA.KEWARGANEGARAAN_AYAH)
    pendidikan_ayah_cell = sheet_obj.cell(row=46, column=4)
    pendidikan_ayah_cell.value = str(self.ORANG_TUA.JENJANG_PENDIDIKAN_AYAH)
    pekerjaan_ayah_cell = sheet_obj.cell(row=47, column=4)
    pekerjaan_ayah_cell.value = str(self.ORANG_TUA.PEKERJAAN_AYAH)
    pengeluaran_ayah_cell = sheet_obj.cell(row=48, column=4)
    pengeluaran_ayah_cell.value = str(self.ORANG_TUA.PENGELUARAN_PER_BULAN_AYAH)
    alamat_ayah_cell = sheet_obj.cell(row=49, column=4)
    alamat_ayah_cell.value = str(self.ORANG_TUA.ALAMAT_AYAH)
    hidup_ayah_cell = sheet_obj.cell(row=50, column=4)
    hidup_ayah_cell.value = str(self.ORANG_TUA.MASIH_HIDUP_AYAH)

    nama_ibu_cell = sheet_obj.cell(row=53, column=4)
    nama_ibu_cell.value = str(self.ORANG_TUA.NAMA_IBU)
    ttl_ibu_cell = sheet_obj.cell(row=54, column=4)
    ttl_ibu_cell.value = str(
        str(self.ORANG_TUA.TEMPAT_LAHIR_IBU)
        + ", "
        + str(self.ORANG_TUA.TAHUN_LAHIR_IBU)
    )
    agama_ibu_cell = sheet_obj.cell(row=55, column=4)
    agama_ibu_cell.value = str(self.ORANG_TUA.AGAMA_IBU)
    kewarganegaraan_ibu_cell = sheet_obj.cell(row=56, column=4)
    kewarganegaraan_ibu_cell.value = str(self.ORANG_TUA.KEWARGANEGARAAN_IBU)
    pendidikan_ibu_cell = sheet_obj.cell(row=57, column=4)
    pendidikan_ibu_cell.value = str(self.ORANG_TUA.JENJANG_PENDIDIKAN_IBU)
    pekerjaan_ibu_cell = sheet_obj.cell(row=58, column=4)
    pekerjaan_ibu_cell.value = str(self.ORANG_TUA.PEKERJAAN_IBU)
    pengeluaran_ibu_cell = sheet_obj.cell(row=59, column=4)
    pengeluaran_ibu_cell.value = str(self.ORANG_TUA.PENGELUARAN_PER_BULAN_IBU)
    alamat_ibu_cell = sheet_obj.cell(row=60, column=4)
    alamat_ibu_cell.value = str(self.ORANG_TUA.ALAMAT_IBU)
    hidup_ibu_cell = sheet_obj.cell(row=61, column=4)
    hidup_ibu_cell.value = str(self.ORANG_TUA.MASIH_HIDUP_IBU)

    nama_wali_cell = sheet_obj.cell(row=64, column=4)
    nama_wali_cell.value = str(self.ORANG_TUA.NAMA_WALI)
    ttl_wali_cell = sheet_obj.cell(row=65, column=4)
    ttl_wali_cell.value = str(
        str(self.ORANG_TUA.TEMPAT_LAHIR_WALI)
        + ", "
        + str(self.ORANG_TUA.TAHUN_LAHIR_WALI)
    )
    agama_wali_cell = sheet_obj.cell(row=66, column=4)
    agama_wali_cell.value = str(self.ORANG_TUA.AGAMA_WALI)
    kewarganegaraan_wali_cell = sheet_obj.cell(row=67, column=4)
    kewarganegaraan_wali_cell.value = str(self.ORANG_TUA.KEWARGANEGARAAN_WALI)
    pendidikan_wali_cell = sheet_obj.cell(row=68, column=4)
    pendidikan_wali_cell.value = str(self.ORANG_TUA.JENJANG_PENDIDIKAN_WALI)
    pekerjaan_wali_cell = sheet_obj.cell(row=69, column=4)
    pekerjaan_wali_cell.value = str(self.ORANG_TUA.PEKERJAAN_WALI)
    pengeluaran_wali_cell = sheet_obj.cell(row=70, column=4)
    pengeluaran_wali_cell.value = str(self.ORANG_TUA.PENGELUARAN_PER_BULAN_WALI)
    alamat_wali_cell = sheet_obj.cell(row=71, column=4)
    alamat_wali_cell.value = str(self.ORANG_TUA.ALAMAT_WALI)

    kesenian_cell = sheet_obj.cell(row=74, column=4)
    kesenian_cell.value = str(self.KESENIAN)
    olahraga_cell = sheet_obj.cell(row=75, column=4)
    olahraga_cell.value = str(self.OLAHRAGA)
    kemasyarakatan_cell = sheet_obj.cell(row=76, column=4)
    kemasyarakatan_cell.value = str(self.KEMASYARAKATAN)
    lain_cell = sheet_obj.cell(row=74, column=4)
    lain_cell.value = str(self.LAIN_LAIN)

    data_bea = []
    for data in beasiswa:
        data_bea.append(
            {"TAHUN": str(data.TAHUN), "KELAS": str(data.KELAS), "DARI": str(data.DARI)}
        )
    db = pd.DataFrame(data_bea)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine="openpyxl")
    db.to_excel(
        writer,
        sheet_name="Beasiswa",
        index=False,
    )
    worksheet2 = writer.sheets["Beasiswa"]

    wb_obj.active = wb_obj["Beasiswa"]
    sheet2_obj = wb_obj.active

    start_row = 0
    start_col = 0

    kelas_row = start_row - 1
    kelas_col = start_col + 2

    mr = worksheet2.max_row + 1
    mc = worksheet2.max_column + 1

    no = 1

    for i in range(2, mr):
        for j in range(1, mc):
            # reading cell value from source excel file
            c = worksheet2.cell(row=i, column=j)

            # writing the read value to destination excel file
            cell = sheet2_obj.cell(row=i + start_row, column=start_col + j)
            cell.value = c.value

            # if c.value is not None and j == 1:
            #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
            #     cell.value = no
            #     no += 1

    if nilai_x_1 is not None:
        n_x_1 = []
        for data in nilai_x_1:
            print(data.MATA_PELAJARAN)
            n_x_1.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        dc = pd.DataFrame(n_x_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dc.to_excel(
            writer,
            sheet_name="RAPORT_X_I",
            index=False,
        )
        worksheet3 = writer.sheets["RAPORT_X_I"]

        wb_obj.active = wb_obj["RAPORT_X_I"]
        sheet3_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet3.max_row + 1
        mc = worksheet3.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet3.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet3_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_x_1 = []
        for data in ekskul_x_1:
            e_x_1.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        dd = pd.DataFrame(e_x_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dd.to_excel(
            writer,
            sheet_name="EKSKUL_X_I",
            index=False,
        )
        worksheet4 = writer.sheets["EKSKUL_X_I"]

        wb_obj.active = wb_obj["EKSKUL_X_I"]
        sheet4_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet4.max_row + 1
        mc = worksheet4.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet4.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet4_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    if nilai_x_2 is not None:
        n_x_2 = []
        for data in nilai_x_2:
            print(data.MATA_PELAJARAN)
            n_x_2.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        de = pd.DataFrame(n_x_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        de.to_excel(
            writer,
            sheet_name="RAPORT_X_II",
            index=False,
        )
        worksheet5 = writer.sheets["RAPORT_X_II"]

        wb_obj.active = wb_obj["RAPORT_X_II"]
        sheet5_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet5.max_row + 1
        mc = worksheet5.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet5.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet5_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_x_2 = []
        for data in ekskul_x_2:
            e_x_2.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        df = pd.DataFrame(e_x_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        df.to_excel(
            writer,
            sheet_name="EKSKUL_X_II",
            index=False,
        )
        worksheet6 = writer.sheets["EKSKUL_X_II"]

        wb_obj.active = wb_obj["EKSKUL_X_II"]
        sheet6_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet6.max_row + 1
        mc = worksheet6.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet6.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet6_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    if nilai_xi_1 is not None:
        n_xi_1 = []
        for data in nilai_xi_1:
            print(data.MATA_PELAJARAN)
            n_xi_1.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        dg = pd.DataFrame(n_xi_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dg.to_excel(
            writer,
            sheet_name="RAPORT_XI_I",
            index=False,
        )
        worksheet7 = writer.sheets["RAPORT_XI_I"]

        wb_obj.active = wb_obj["RAPORT_XI_I"]
        sheet7_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet7.max_row + 1
        mc = worksheet7.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet7.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet7_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_xi_1 = []
        for data in ekskul_xi_1:
            e_xi_1.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        dh = pd.DataFrame(e_xi_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dh.to_excel(
            writer,
            sheet_name="EKSKUL_XI_I",
            index=False,
        )
        worksheet8 = writer.sheets["EKSKUL_XI_I"]

        wb_obj.active = wb_obj["EKSKUL_XI_I"]
        sheet8_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet8.max_row + 1
        mc = worksheet8.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet8.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet8_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    if nilai_xi_2 is not None:

        n_xi_2 = []
        for data in nilai_xi_2:
            print(data.MATA_PELAJARAN)
            n_xi_2.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        di = pd.DataFrame(n_xi_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        di.to_excel(
            writer,
            sheet_name="RAPORT_XI_II",
            index=False,
        )
        worksheet9 = writer.sheets["RAPORT_XI_II"]

        wb_obj.active = wb_obj["RAPORT_XI_II"]
        sheet9_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet9.max_row + 1
        mc = worksheet9.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet9.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet9_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_xi_2 = []
        for data in ekskul_xi_2:
            e_xi_2.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        dj = pd.DataFrame(e_xi_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dj.to_excel(
            writer,
            sheet_name="EKSKUL_XI_II",
            index=False,
        )
        worksheet10 = writer.sheets["EKSKUL_XI_II"]

        wb_obj.active = wb_obj["EKSKUL_XI_II"]
        sheet10_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet10.max_row + 1
        mc = worksheet10.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet10.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet10_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    if nilai_xii_1 is not None:
        n_xii_1 = []
        for data in nilai_xii_1:
            print(data.MATA_PELAJARAN)
            n_xii_1.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        dk = pd.DataFrame(n_xii_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dk.to_excel(
            writer,
            sheet_name="RAPORT_XII_I",
            index=False,
        )
        worksheet11 = writer.sheets["RAPORT_XII_I"]

        wb_obj.active = wb_obj["RAPORT_XII_I"]
        sheet11_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet11.max_row + 1
        mc = worksheet11.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet11.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet11_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_xii_1 = []
        for data in ekskul_xii_1:
            e_xii_1.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        dl = pd.DataFrame(e_xii_1)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dl.to_excel(
            writer,
            sheet_name="EKSKUL_XII_I",
            index=False,
        )
        worksheet12 = writer.sheets["EKSKUL_XII_I"]

        wb_obj.active = wb_obj["EKSKUL_XII_I"]
        sheet12_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet12.max_row + 1
        mc = worksheet12.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet12.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet12_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    if nilai_xii_2 is not None:
        n_xii_2 = []
        for data in nilai_xii_2:
            print(data.MATA_PELAJARAN)
            n_xii_2.append(
                {
                    "MATA_PELAJARAN": data.MATA_PELAJARAN,
                    "BEBAN": data.BEBAN,
                    "NILAI_PENGETAHUAN": data.NILAI_PENGETAHUAN,
                    "NILAI_KETERAMPILAN": data.NILAI_KETERAMPILAN,
                    "DESKRIPSI_PENGETAHUAN": data.DESKRIPSI_PENGETAHUAN,
                    "DESKRIPSI_KETERAMPILAN": data.DESKRIPSI_KETERAMPILAN,
                }
            )
        dm = pd.DataFrame(n_xii_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dm.to_excel(
            writer,
            sheet_name="RAPORT_XII_II",
            index=False,
        )
        worksheet13 = writer.sheets["RAPORT_XII_II"]

        wb_obj.active = wb_obj["RAPORT_XII_II"]
        sheet13_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet13.max_row + 1
        mc = worksheet13.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet13.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet13_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

        e_xii_2 = []
        for data in ekskul_xii_2:
            e_xii_2.append(
                {
                    "EKSKUL": data.DATA_ANGGOTA.EKSKUL,
                    "PREDIKAT": data.PREDIKAT,
                    "DESKRIPSI": data.DESKRIPSI,
                }
            )

        dn = pd.DataFrame(e_xii_2)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine="openpyxl")
        dn.to_excel(
            writer,
            sheet_name="EKSKUL_XII_II",
            index=False,
        )
        worksheet14 = writer.sheets["EKSKUL_XII_II"]

        wb_obj.active = wb_obj["EKSKUL_XII_II"]
        sheet14_obj = wb_obj.active

        start_row = 0
        start_col = 0

        kelas_row = start_row - 1
        kelas_col = start_col + 2

        mr = worksheet14.max_row + 1
        mc = worksheet14.max_column + 1

        no = 1

        for i in range(2, mr):
            for j in range(1, mc):
                # reading cell value from source excel file
                c = worksheet14.cell(row=i, column=j)

                # writing the read value to destination excel file
                cell = sheet14_obj.cell(row=i + start_row, column=start_col + j)
                cell.value = c.value

                # if c.value is not None and j == 1:
                #     cell = sheet2_obj.cell(row = i + kelas_row + 2, column = 1)
                #     cell.value = no
                #     no += 1

    wb_obj.active = wb_obj["Data_Siswa"]
    virtual_workbook = BytesIO()
    wb_obj.save(virtual_workbook)
    self.HASIL_EXPORT = ContentFile(
        virtual_workbook.getvalue(), str(self.NIS) + str(self.ID) + ".xlsx"
    )
