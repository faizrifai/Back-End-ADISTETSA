from email.mime import image
from io import  BytesIO
from django.core.files.base import ContentFile
import openpyxl
import pandas as pd
import pdfkit
from win32com import client
from pdfrw import pdfwriter
from xlsx2html import xlsx2html
from fpdf import FPDF

def buat_kuitansi(self): 
   # wb_obj = openpyxl.load_workbook(self.TEMPLATE.path)
   # sheet_obj = wb_obj.active
   # nama_cell =  sheet_obj.cell(row = 8, column= 4)
   # nama_cell.value = self.NAMA_SISWA.NIS.NAMA
   # kelas_cell =  sheet_obj.cell(row = 9, column= 4)
   # kelas_cell.value = str(self.NAMA_SISWA.KELAS)
   # nama_cell =  sheet_obj.cell(row = 10, column= 4)
   # nama_cell.value = self.NAMA_SISWA.NIS.NIS
   # total_pembayaran = 0
   
   # print ('test' + self.PEMBAYARAN_DPSM_INSINDENTAL)
   
   # if self.NOMINAL_SPP != '' :
   #    bulan = self.PEMBAYARAN_SPP.split(',')
   #    jenis_pembayaran_cell =  sheet_obj.cell(row = 11, column= 3)
   #    jenis_pembayaran_cell.value = 'X' 
   #    total_pembayaran += (len(bulan)*int(self.NOMINAL_SPP)) 
   #    print (len(bulan))
   #    for i in range(len(bulan)):
   #       cur = bulan[i-1].strip()
   #       print(cur)
   #       if cur in 'Januari':
   #          bulan_cell =  sheet_obj.cell(row = 6, column= 8)
   #          bulan_cell.value = 'X'
   #       elif cur in 'Februari' :
   #          bulan_cell =  sheet_obj.cell(row = 7, column= 8)
   #          bulan_cell.value = 'X'
   #       elif cur in 'Maret':
   #          bulan_cell =  sheet_obj.cell(row = 8, column= 8)
   #          bulan_cell.value = 'X'
   #       elif cur in 'April':
   #          bulan_cell =  sheet_obj.cell(row = 9, column= 8)
   #          bulan_cell.value = 'X'
   #       elif cur in 'Mei':
   #          bulan_cell =  sheet_obj.cell(row = 10, column= 8)
   #          bulan_cell.value = 'X'   
   #       elif cur in 'Juni':
   #          bulan_cell =  sheet_obj.cell(row = 10, column= 8)
   #          bulan_cell.value = 'X'   
   #       elif cur in 'Juli':
   #          bulan_cell =  sheet_obj.cell(row = 6, column= 6)
   #          bulan_cell.value = 'X'
   #       elif cur in 'Agustus':
   #          bulan_cell =  sheet_obj.cell(row = 7, column= 6)
   #          bulan_cell.value = 'X'
   #       elif cur in 'September':
   #          bulan_cell =  sheet_obj.cell(row = 8, column= 6)
   #          bulan_cell.value = 'X'
   #       elif cur in 'Oktober':
   #          bulan_cell =  sheet_obj.cell(row = 9, column= 6)
   #          bulan_cell.value = 'X'
   #       elif cur in 'November':
   #          bulan_cell =  sheet_obj.cell(row = 10, column= 6)
   #          bulan_cell.value = 'X'   
   #       elif cur in 'Desember':
   #          bulan_cell =  sheet_obj.cell(row = 11, column= 6)
   #          bulan_cell.value = 'X'    
      
   # if self.PEMBAYARAN_DPSM_RUTIN != '' :
   #    jenis_pembayaran_cell =  sheet_obj.cell(row = 12, column= 3)
   #    jenis_pembayaran_cell.value = 'X'
   #    total_pembayaran += int(float(self.PEMBAYARAN_DPSM_RUTIN))
      
   # if self.PEMBAYARAN_DPSM_INSINDENTAL != 0 :
   #    jenis_pembayaran_cell =  sheet_obj.cell(row = 13, column= 3)
   #    jenis_pembayaran_cell.value = 'X'
   #    value = self.PEMBAYARAN_DPSM_INSINDENTAL
   #    print (value)
   #    total_pembayaran += int(value)
      
   # elif self.BIMBEL != 0 :
   #    jenis_pembayaran_cell =  sheet_obj.cell(row = 14, column= 3)
   #    jenis_pembayaran_cell.value = 'X'
   #    total_pembayaran += int(self.BIMBEL)
      
   # nominal_cell =  sheet_obj.cell(row = 15, column= 4)
   # nominal_cell.value = 'Rp. ' + str(total_pembayaran) 
   # virtual_workbook = BytesIO()
   # wb_obj.save(virtual_workbook)
   
   DPSMR = ''
   DPSMI = ''
   BIMBEL = ''
   SPP = ''
   
   NILAI_SPP = ''
   
   JANUARI = ''
   FEBRUARI = ''
   MARET = ''
   APRIL = ''
   MEI = ''
   JUNI = ''
   JULI = ''
   AGUSTUS = ''
   SEPTEMBER = ''
   OKTOBER = ''
   NOVEMBER = ''
   DESEMBER = ''
   total_pembayaran = 0
   
   if self.NOMINAL_SPP != '0' :
      bulan = self.PEMBAYARAN_SPP.split(',')
      SPP = 'X'
      total_pembayaran += (len(bulan)*int(self.NOMINAL_SPP)) 
      print (len(bulan))
      for i in range(len(bulan)):
         cur = bulan[i-1].strip()
         print(cur)
         if cur in 'Januari':
            JANUARI = 'X'
         elif cur in 'Februari' :
            FEBRUARI = 'X'
         elif cur in 'Maret':
            MARET = 'X'
         elif cur in 'April':
            APRIL = 'X'
         elif cur in 'Mei':
            MEI = ''
         elif cur in 'Juni':
            JUNI = 'X'
         elif cur in 'Juli':
            JULI = 'X'
         elif cur in 'Agustus':
            AGUSTUS = 'X'
         elif cur in 'September':
            SEPTEMBER = 'X'
         elif cur in 'Oktober':
            OKTOBER = 'X'
         elif cur in 'November':
            NOVEMBER = 'X'
         elif cur in 'Desember':
            DESEMBER = 'X'
      
   if self.PEMBAYARAN_DPSM_RUTIN != '0' :
      DPSMR = 'X'
      total_pembayaran += int(float(self.PEMBAYARAN_DPSM_RUTIN))
      
   if self.PEMBAYARAN_DPSM_INSINDENTAL != '0' :
      DPSMI = 'X'
      value = self.PEMBAYARAN_DPSM_INSINDENTAL
      print (value)
      total_pembayaran += int(value)
      
   if self.BIMBEL != '0' :
      BIMBEL = 'X'
      total_pembayaran += int(self.BIMBEL)
   
   pdf = FPDF()
  
   # Add a page
   pdf.add_page()
   
   # set style and size of font 
   # that you want in the pdf
   
   pdf.set_font("Times", size = 15)
   # create a cell
   # pdf.cell(200, 10, )
   pdf.image("Logo.png", x = None, y = None, w = 190, h = 35, type = '', link = '')
   pdf.cell(200, 12, txt = "KUITANSI PEMBAYARAN", 
            ln = 1, align = 'C')
   pdf.line(10, 55, 200, 55)
   pdf.line(10, 56, 200, 56)
   pdf.set_font("Times", size = 15)
   # add another cell
   pdf.cell(200, 10, txt = str('NAMA SISWA    = ' + self.NAMA_SISWA.NIS.NAMA),
            ln = 2, align = 'L')
   pdf.cell(200, 10, txt = str('KELAS                 = ' + str(self.NAMA_SISWA.KELAS)),
            ln = 2, align = 'L')
   pdf.cell(200, 10, txt = str('NOMOR INDUK = ' + str(self.NAMA_SISWA.NIS.NIS)),
            ln = 2, align = 'L')
   pdf.cell(200, 10, txt = 'JENIS PEMBAYARAN :',
            ln = 2, align = 'L')
   pdf.set_font("Times", size = 15)
   pdf.set_xy(10, 100)
   # pdf.ln(1)
   pdf.multi_cell(20, 10, txt= DPSMR, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(30, 100)
   # pdf.ln(1)
   pdf.multi_cell(170, 10, txt= 'Dana Peran Serta Masyarakat Rutin', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 110)
   pdf.multi_cell(20, 10, txt= DPSMI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(30, 110)
   # pdf.ln(1)
   pdf.multi_cell(170, 10, txt= 'Dana Peran Serta Masyarakat Insindental', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 120)
   pdf.multi_cell(20, 10, txt= BIMBEL, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(30, 120)
   # pdf.ln(1)
   pdf.multi_cell(170, 10, txt= 'Bimbel', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 130)
   pdf.multi_cell(20, 10, txt= SPP, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(30, 130)
   # pdf.ln(1)
   pdf.multi_cell(170, 10, txt= 'SPP', border = 1, 
                align= 'L', fill=bool)
   pdf.set_font("Times", size = 15)
   pdf.set_xy(10, 140)
   pdf.multi_cell(75, 10, txt= 'JUMLAH SPP PER BULAN = ', border = 0, 
                align= 'L', fill=bool)
   pdf.set_xy(85, 140)
   # pdf.ln(1)
   pdf.multi_cell(125, 10, txt= self.NOMINAL_SPP, border = 0, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 150)
   pdf.multi_cell(80, 10, txt= 'PEMBAYARAN BULAN : ', border = 0, 
                align= 'L', fill=bool)
   pdf.line(10, 161, 200, 161)
   pdf.set_font("Times", size = 15)
   pdf.set_xy(10, 165)
   pdf.multi_cell(10, 10, txt= JANUARI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(20, 165)
   pdf.multi_cell(50, 10, txt= 'JANUARI', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(70, 165)
   pdf.multi_cell(10, 10, txt= FEBRUARI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(80, 165)
   pdf.multi_cell(50, 10, txt= 'FEBRUARI', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(130, 165)
   pdf.multi_cell(10, 10, txt= MARET, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(140, 165)
   pdf.multi_cell(60, 10, txt= 'MARET', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 175)
   pdf.multi_cell(10, 10, txt= APRIL, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(20, 175)
   pdf.multi_cell(50, 10, txt= 'APRIL', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(70, 175)
   pdf.multi_cell(10, 10, txt= MEI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(80, 175)
   pdf.multi_cell(50, 10, txt= 'MEI', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(130, 175)
   pdf.multi_cell(10, 10, txt= JUNI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(140, 175)
   pdf.multi_cell(60, 10, txt= 'JUNI', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 185)
   pdf.multi_cell(10, 10, txt= JULI, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(20, 185)
   pdf.multi_cell(50, 10, txt= 'JULI', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(70, 185)
   pdf.multi_cell(10, 10, txt= AGUSTUS, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(80, 185)
   pdf.multi_cell(50, 10, txt= 'AGUSTUS', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(130, 185)
   pdf.multi_cell(10, 10, txt= SEPTEMBER, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(140, 185)
   pdf.multi_cell(60, 10, txt= 'SEPTEMBER', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(10, 195)
   pdf.multi_cell(10, 10, txt= OKTOBER, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(20, 195)
   pdf.multi_cell(50, 10, txt= 'OKTOBER', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(70, 195)
   pdf.multi_cell(10, 10, txt= NOVEMBER, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(80, 195)
   pdf.multi_cell(50, 10, txt= 'NOVEMBER', border = 1, 
                align= 'L', fill=bool)
   pdf.set_xy(130, 195)
   pdf.multi_cell(10, 10, txt= DESEMBER, border = 1, 
                align= 'C', fill=bool)
   pdf.set_xy(140, 195)
   pdf.multi_cell(60, 10, txt= 'DESEMBER', border = 1, 
                align= 'L', fill=bool)
   pdf.line(10, 207, 200, 207)
   pdf.line(10, 208, 200, 208)
   pdf.set_font("Times", size = 15)
   pdf.set_xy(10, 210)
   pdf.multi_cell(170, 5, txt= str('TOTAL PEMBAYARAN =  Rp.' + str(total_pembayaran)), border = 0, 
                align= 'L', fill=bool)
   pdf.set_xy(100, 215)
   pdf.multi_cell(90, 10, txt= str('Malang, ' + str(self.TANGGAL_PEMBAYARAN)), border = 0, 
                align= 'R', fill=bool)
   pdf.set_xy(10, 225)
   pdf.multi_cell(90, 5, txt= 'Petugas / Penerima', border = 0, 
                align= 'C', fill=bool)
   pdf.set_xy(110, 225)
   pdf.multi_cell(90, 5, txt= 'Penyetor', border = 0, 
                align= 'C', fill=bool)
   pdf.set_xy(10, 267)
   pdf.multi_cell(90, 5, txt= '.................', border = 0, 
                align= 'C', fill=bool)
   pdf.set_xy(110, 267)
   pdf.multi_cell(90, 5, txt= '.................', border = 0, 
                align= 'C', fill=bool)
   pdf.set_font("Times", size = 8)
   pdf.set_xy(10, 275)
   pdf.multi_cell(90, 1, txt= '*Mohon slip bukti disimpan dengan baik', border = 0, 
                align= 'L', fill=bool)
   
   # save the pdf with name .pdf
   # pdf.output("GFG.pdf")   
   byte_string = pdf.output(dest='S').encode('latin-1')  # Probably what you want
   stream = BytesIO(byte_string) 

   # out_stream = xlsx2html(self.TEMPLATE.path)
   # out_stream.seek(0)
   # print(out_stream.read())
   
   # xlsx2html(self.TEMPLATE.path, 'output.html')
   
   # pw = pdfwriter()
   # pw.setFont('Courier', 12)
   # pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
   # pw.setFooter('Generated using openpyxl and xtopdf')

   # ws_range = sheet_obj.iter_rows('A1:H13')
   # for row in ws_range:
   #    s = ''
   #    for cell in row:
   #       if cell.value is None:
   #             s += ' ' * 11
   #       else:
   #             s += str(cell.value).rjust(10) + ' '
   #    pw.writeLine(s)
   # pw.savePage()
   # pw.close()

   # options =  HtmlSaveOptions()
   # # show tooltips
   # options.setAddTooltipText(True)
   # # save workbook as HTML file
   # wb.save("workbook.html", options)
   # wb_obj.save("workbook.html")
   # df = pd.read_excel(virtual_workbook.getvalue())
   # df.to_html("file.html")
   # pdfkit.from_file("file.html", "file.pdf")
   # Open Microsoft Excel
   # excel = client.Dispatch("Excel.Application")
   
   # # Read Excel File
   # sheets = excel.Workbooks.Open('Excel File Path')
   # work_sheets = sheets.Worksheets[0]
   
   # # Convert into PDF File
   # work_sheets.ExportAsFixedFormat(0, 'PDF File Path')
   
   return ContentFile(stream.getvalue(), 'kuitansi_' + str(self.NAMA_SISWA.NIS.NAMA) + '_'+ str(self.TANGGAL_PEMBAYARAN)+'_'+ str(self.ID) + '.pdf')
        
   
