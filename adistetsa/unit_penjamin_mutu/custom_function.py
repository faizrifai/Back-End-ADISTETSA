from asyncore import write
from genericpath import exists
from heapq import merge
import pandas as pd
from io import BytesIO
from django.core.files.base import ContentFile
from django.db.models import Count
from kurikulum.models import JadwalMengajar, OfferingKelas, KelasSiswa, Kelas
from dataprofil.models import DataRiwayatKepangkatanGuru
from utility.custom_function import gabung_dictionary
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
import datetime

import time

def value_in_list_of_dict(key, value, d):
    for data in d:
        if data[key] == value:
            return True
        
    return False

def multiple_value_exist(v: dict, d: dict):
    keys = v.keys()
    for data in d:
        kebenaran = True
        for key in keys:
            value = v[key]
        
            if data[key] == value:
                kebenaran = kebenaran and True
            else:
                kebenaran = kebenaran and False
                
        if kebenaran:
            return True
            
    return False


def get_column(df, column_name):
    col_no = chr(65 + df.columns.get_loc(column_name))
    col_no = col_no + ':' + col_no
    
    return col_no

def get_columns_from_worksheet(ws):
  return {
      cell.value: {
          'letter': openpyxl.utils.get_column_letter(cell.column),
          'number': cell.column - 1
      } for cell in ws[1] if cell.value
  }

# def apply_style_to_cell(worksheet, width, font, border):
#     column_letter = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(worksheet.max_column))
#     for letter in column_letter:
#         worksheet.column_dimensions[letter].width = width
#         for cell in worksheet[letter + ':' + letter]:
#             cell.font = font
#             cell.border = border
def apply_style_to_cell(worksheet, width, font):
    column_letter = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(worksheet.max_column))
    for letter in column_letter:
        worksheet.column_dimensions[letter].width = width
        for cell in worksheet[letter + ':' + letter]:
            cell.font = font
            
def merge_cell(worksheet, dataframe, col_name, col_merge):
    col = get_columns_from_worksheet(worksheet)
    col_dif = col[col_merge]['number'] + 1
    
    for car in dataframe[col_name].unique():
        # find indices and add one to account for header
        u=dataframe.loc[dataframe[col_name]==car].index.values + 1

        if len(u) < 2:
            pass # do not merge cells if there is only one car name
        else:
            worksheet.merge_cells(start_row=u[0] + 1, start_column=col_dif, end_row=u[-1] + 1, end_column=col_dif)
            
def cari_pangkat(self):
    pangkat = DataRiwayatKepangkatanGuru.objects.all()
    for data in pangkat:
        if self == data.OWNER:
            return str(data.PANGKAT_GOLONGAN)
        
    return '-'

def pembagian_jadwal_mengajar(self):
    start = time.time()

    kelas = OfferingKelas.objects.all().order_by('KELAS__TINGKATAN', '-KELAS__JURUSAN__NAMA', 'OFFERING')
    jadwal = JadwalMengajar.objects.all().filter(TAHUN_AJARAN=self.TAHUN_AJARAN, SEMESTER = self.SEMESTER).order_by('GURU').annotate(JUMLAH=Count('WAKTU_PELAJARAN'))
    kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
    
    guru = []
    for data in jadwal:
        exist = multiple_value_exist({
            'GURU': data.GURU.NAMA_LENGKAP,
            'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
        }, guru)

        if not (exist):
            guru.append({
                'GURU': data.GURU.NAMA_LENGKAP,
                'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
            })

    jumlah = {}

    for data in kelas:
        jumlah[str(data)] = ''
    
    jumlah['JUMLAH_JAM_ATAU_RASIO'] = 0
    jumlah['JAM_TAMBAHAN'] = ''
    jumlah['TOTAL_JAM'] = 0
    jumlah['TUGAS_TAMBAHAN'] = ''
    
    # pembagian jadwal mengajar
    jumlah_keseluruhan = []
    for g in guru:
        if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru, jumlah)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'])

        for data in jadwal:
            if data.GURU.NAMA_LENGKAP == g['GURU'] and str(data.MATA_PELAJARAN) == g['MATA_PELAJARAN']:
                data_baru['JUMLAH_JAM_ATAU_RASIO'] += data.JUMLAH

                if data_baru[str(data.KELAS)] == '':
                    data_baru[str(data.KELAS)] = data.JUMLAH
                else:
                    data_baru[str(data.KELAS)] += data.JUMLAH
                    
            if data.GURU.NAMA_LENGKAP == g['GURU']:
                data_baru['TOTAL_JAM'] += data.JUMLAH

        if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)


    # total kelas
    jumlah_siswa = {}
    for data in kelas:
        jumlah = kelas_siswa.filter(KELAS = data).count()    
        jumlah_siswa[str(data)] = jumlah
        
    
    for data in kelas:
        jumlah_siswa[str(data.KELAS.TINGKATAN)] = kelas_siswa.filter(KELAS__KELAS__TINGKATAN = data.KELAS.TINGKATAN).count()
    
    jumlah_total = []
    jumlah_total.append(jumlah_siswa)
    
    
    df_ = pd.DataFrame(jumlah_keseluruhan)
    df = pd.DataFrame(jumlah_keseluruhan).set_index(['GURU', 'MATA_PELAJARAN'])
    dt = pd.DataFrame(jumlah_total)
    
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    
    df.to_excel(writer, sheet_name=self.KATEGORI, index=True, )
    dt.to_excel(writer, sheet_name='TOTAL_JUMLAH', index=False,)
    
    worksheet = writer.sheets[self.KATEGORI]
    worksheet2 = writer.sheets['TOTAL_JUMLAH']
    
    # def apply_style_to_cell(worksheet, width, font, border):
    # bd = Side(style='thick', color="000000")
    # apply_style_to_cell(worksheet, 30, 'Times New Roman', Border(left=bd, top=bd, right=bd, bottom=bd))
    apply_style_to_cell(worksheet,20,'Times New Roman')
    apply_style_to_cell(worksheet2,20,'Times New Roman')
    
    # col = get_columns_from_worksheet(worksheet)
    # total_jam = col['TOTAL_JAM']['number'] + 1
    
    merge_cell(worksheet, df_, 'GURU', 'TOTAL_JAM')
    writer.save()
    
    return ContentFile(b.getvalue(), self.KATEGORI + '_' +str (datetime.date.today()) +'.xlsx')

def rekapitulasi_jam_mengajar(self):
    kelas_tingkatan = Kelas.objects.all().order_by('TINGKATAN',)
    jadwal = JadwalMengajar.objects.all().filter(TAHUN_AJARAN=self.TAHUN_AJARAN, SEMESTER = self.SEMESTER).order_by('GURU').annotate(JUMLAH=Count('WAKTU_PELAJARAN'))
    kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
    
    guru = []
    for data in jadwal:
        exist = multiple_value_exist({
            'GURU': data.GURU.NAMA_LENGKAP,
            'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
        }, guru)
        
        if not (exist):
            guru.append({
                'GURU': data.GURU.NAMA_LENGKAP,
                'NIP': data.GURU.NIP,
                'PANGKAT': data.GURU.PANGKAT,
                'GOL': cari_pangkat(data.GURU),
                'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
            })
            
    jumlah = {}
    
    for data in kelas_tingkatan:
        jumlah[str(data.TINGKATAN)] = 0

    jumlah_keseluruhan = []
    for g in guru:
        if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru, jumlah)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'])

        for data in jadwal:
            if data.GURU.NAMA_LENGKAP == g['GURU'] and str(data.MATA_PELAJARAN) == g['MATA_PELAJARAN']:
                tingkatan = data.KELAS.KELAS.TINGKATAN
                data_baru[tingkatan] += data.JUMLAH

        if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)

    df = pd.DataFrame(jumlah_keseluruhan)
    df_ = pd.DataFrame(jumlah_keseluruhan).set_index(['GURU', 'NIP',])
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    df_.to_excel(writer, sheet_name=self.KATEGORI, index=True, )
    print(df_)
    
    worksheet = writer.sheets[self.KATEGORI]
    merge_cell(worksheet, df, 'GURU', 'NIP')
    merge_cell(worksheet, df, 'GURU', 'PANGKAT')
    merge_cell(worksheet, df, 'GURU', 'GOL')    
    apply_style_to_cell(worksheet,20,'Times New Roman')
    
    writer.save()
    return ContentFile(b.getvalue(), self.KATEGORI + '_' +str (datetime.date.today()) +'.xlsx') 

def pembagian_tugas_guru_bk(self, count):
    pembagian_tugas = count 
    kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
    data_pembagian =[]
    data_cek = {}
    guru = []
    for data in pembagian_tugas:
        exist = multiple_value_exist({
            'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
        }, guru)
        
        if not (exist):
            guru.append({
                'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
                'PANGKAT_GOLONGAN': data.DATA_GURU.PANGKAT + ',' + str(cari_pangkat(data.DATA_GURU)),
            })
    
            
    jumlah = {}
    jumlah_keseluruhan = []
    
    for g in guru:
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru, jumlah)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'])
            
        
        list_kelas_x =  []
        list_kelas_xi =  []
        list_kelas_xii =  []
        
        jumlah_kelas_x = []
        jumlah_kelas_xi = []
        jumlah_kelas_xii = []
        # tugas_individual = pembagian_tugas.filter(DAT)
        for data in pembagian_tugas.filter(DATA_GURU__NAMA_LENGKAP = g['NAMA_NIP_JABATAN']):
            for data_kelas in data.DATA_KELAS.all():
                print (data_kelas)
                if data_kelas.KELAS.TINGKATAN == 'X' :
                    print (data_kelas)
                    list_kelas_x.append(str(data_kelas))
                    jumlah_kelas_x.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XI' :
                    list_kelas_xi.append(str(data_kelas))               
                    jumlah_kelas_xi.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XII' :
                    list_kelas_xii.append(str(data_kelas))               
                    jumlah_kelas_xii.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                # jumlah[str(data)]  = kelas_siswa.filter(KELAS = data_kelas,).count()

        # data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        data_baru['JUMLAH_KELAS_X'] = ", ".join(jumlah_kelas_x)
        data_baru['KELAS_XI'] = ", ".join(list_kelas_xi)
        data_baru['JUMLAH_KELAS_XI'] = ", ".join(jumlah_kelas_xi)
        data_baru['KELAS_XII'] = ", ".join(list_kelas_xii)
        data_baru['JUMLAH_KELAS_XII'] = ", ".join(jumlah_kelas_xii)
        data_baru['KETERANGAN'] = data.KETERANGAN_TUGAS
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)
    print (jumlah_keseluruhan)
    dt = pd.DataFrame(jumlah_keseluruhan)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    dt.to_excel(writer, sheet_name='PEMBAGIAN_BK', index=False)
    worksheet = writer.sheets['PEMBAGIAN_BK']    
    apply_style_to_cell(worksheet,20,'Times New Roman')
    writer.save()
    self.FILE = ContentFile(b.getvalue(), 'PEMBAGIAN_BK' + '_' +str (datetime.date.today()) +'.xlsx') 

def pembagian_tugas_guru_tik(self, count):
    pembagian_tugas = count 
    kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
    data_pembagian =[]
    data_cek = {}
    guru = []
    for data in pembagian_tugas:
        exist = multiple_value_exist({
            'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
        }, guru)
        
        if not (exist):
            guru.append({
                'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
                'PANGKAT_GOLONGAN': data.DATA_GURU.PANGKAT + ',' + str(cari_pangkat(data.DATA_GURU)),
            })
    
            
    jumlah = {}
    jumlah_keseluruhan = []
    
    for g in guru:
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru, jumlah)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'])
            
        
        list_kelas_x =  []
        list_kelas_xi =  []
        list_kelas_xii =  []
        
        jumlah_kelas_x = []
        jumlah_kelas_xi = []
        jumlah_kelas_xii = []
        # tugas_individual = pembagian_tugas.filter(DAT)
        for data in pembagian_tugas.filter(DATA_GURU__NAMA_LENGKAP = g['NAMA_NIP_JABATAN']):
            for data_kelas in data.DATA_KELAS.all():
                print (data_kelas)
                if data_kelas.KELAS.TINGKATAN == 'X' :
                    print (data_kelas)
                    list_kelas_x.append(str(data_kelas))
                    jumlah_kelas_x.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XI' :
                    list_kelas_xi.append(str(data_kelas))               
                    jumlah_kelas_xi.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XII' :
                    list_kelas_xii.append(str(data_kelas))               
                    jumlah_kelas_xii.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                # jumlah[str(data)]  = kelas_siswa.filter(KELAS = data_kelas,).count()

        # data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        data_baru['JUMLAH_KELAS_X'] = ", ".join(jumlah_kelas_x)
        data_baru['KELAS_XI'] = ", ".join(list_kelas_xi)
        data_baru['JUMLAH_KELAS_XI'] = ", ".join(jumlah_kelas_xi)
        data_baru['KELAS_XII'] = ", ".join(list_kelas_xii)
        data_baru['JUMLAH_KELAS_XII'] = ", ".join(jumlah_kelas_xii)
        data_baru['KETERANGAN'] = data.KETERANGAN_TUGAS
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)
    print (jumlah_keseluruhan)
    dt = pd.DataFrame(jumlah_keseluruhan)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    dt.to_excel(writer, sheet_name='PEMBAGIAN_TIK', index=False)
    worksheet = writer.sheets['PEMBAGIAN_TIK'] 
    apply_style_to_cell(worksheet,20,'Times New Roman')
    writer.save()   
    self.FILE = ContentFile(b.getvalue(), 'PEMBAGIAN_TIK' + '_' +str (datetime.date.today()) +'.xlsx') 

def pembagian_tugas_pokok_tendik(self, count):
    pembagian_tugas = count
    guru = []
    for data in pembagian_tugas.filter(TAHUN_AJARAN = self.TAHUN_AJARAN):
        exist = multiple_value_exist({
            'NAMA' : data.DATA_GURU.NAMA_LENGKAP
        }, guru)
        if not (exist):
            guru.append({
                'NAMA': data.DATA_GURU.NAMA_LENGKAP,
                'NIP' : data.DATA_GURU.NIP, 
                'Gol' : cari_pangkat(data.DATA_GURU),
                'TUGAS_POKOK' : data.TUGAS_POKOK,
                'TUGAS_TAMBAHAN' : data.TUGAS_TAMBAHAN,
            })
    jumlah_keseluruhan = []
    for g in guru:
        if not any(d['NAMA'] == g['NAMA'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['NAMA'] == g['NAMA'])
        if not any(d['NAMA'] == g['NAMA'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)
    df = pd.DataFrame(jumlah_keseluruhan)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    df.to_excel(writer, sheet_name='TUGAS_POKOK', index=False, )
    worksheet = writer.sheets['TUGAS_POKOK']
    apply_style_to_cell(worksheet,30,'Times New Roman')
    writer.save()
    self.FILE = ContentFile(b.getvalue(), 'PEMBAGIAN_TUGAS' +'_' +str (datetime.date.today()) + '.xlsx') 

def rincian_tugas_pokok_tendik(self, count):
    pembagian_tugas = count
    guru = []
    for data in pembagian_tugas.filter(TAHUN_AJARAN = self.TAHUN_AJARAN):
        exist = multiple_value_exist({
            'NAMA_JABATAN' : str(data.DATA_GURU.NAMA_LENGKAP) +', ' + str(data.DATA_GURU.PANGKAT), 
        }, guru)
        if not (exist):
            guru.append({
                'NAMA_JABATAN' : str(data.DATA_GURU.NAMA_LENGKAP) +', ' + str(data.DATA_GURU.PANGKAT),
                'TUGAS_POKOK' : data.TUGAS_POKOK,
                'TUGAS_TAMBAHAN' : data.TUGAS_TAMBAHAN,
            })
    jumlah_keseluruhan = []
    for g in guru:
        if not any(d['NAMA_JABATAN'] == g['NAMA_JABATAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['NAMA_JABATAN'] == g['NAMA_JABATAN'])
        if not any(d['NAMA_JABATAN'] == g['NAMA_JABATAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)
    df = pd.DataFrame(jumlah_keseluruhan)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    df.to_excel(writer, sheet_name='RINCIAN_TUGAS', index=False, )
    worksheet = writer.sheets['RINCIAN_TUGAS']
    apply_style_to_cell(worksheet,40,'Times New Roman')
    writer.save()
    self.FILE = ContentFile(b.getvalue(), 'RINCIAN_TUGAS' + '_' +str (datetime.date.today()) +'.xlsx')

def tugas_tambahan_kepanitiaan(self, count):
    pembagian_tugas = count
    guru = []
    for data in pembagian_tugas.filter(TAHUN_AJARAN = self.TAHUN_AJARAN):
        guru.append({
            'KODE_BIDANG' : data.BIDANG.KODE_BIDANG,
            'BIDANG' : data.BIDANG.NAMA_BIDANG,
            'NAMA' : data.DATA_GURU.NIP,
            'Gol' : cari_pangkat(data.DATA_GURU),
            'PANGKAT' : data.DATA_GURU.PANGKAT,
            'TUGAS' : data.TUGAS,
            'SUB_BIDANG' : data.SUB_BIDANG
        })
    df = pd.DataFrame(guru).set_index(['BIDANG', 'KODE_BIDANG'])
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    df.to_excel(writer, sheet_name='TUGAS_TAMBAHAN_KEPANITIAAN', index=True, )
    worksheet = writer.sheets['TUGAS_TAMBAHAN_KEPANITIAAN']
    apply_style_to_cell(worksheet,20,'Times New Roman')
    
    writer.save()
    self.FILE = ContentFile(b.getvalue(), 'TUGAS_TAMBAHAN_KEPANITIAAN' +'_' +str (datetime.date.today()) + '.xlsx')
    
def buat_file_prototype(self):
    if self.KATEGORI == 'Pembagian Jadwal Mengajar':
        # check execution time -- start
        start = time.time()

        kelas = OfferingKelas.objects.all().order_by('KELAS__TINGKATAN', '-KELAS__JURUSAN__NAMA', 'OFFERING')
        jadwal = JadwalMengajar.objects.all().filter(TAHUN_AJARAN=self.TAHUN_AJARAN, SEMESTER = self.SEMESTER).order_by('GURU').annotate(JUMLAH=Count('WAKTU_PELAJARAN'))
        kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
        
        guru = []
        for data in jadwal:
            exist = multiple_value_exist({
                'GURU': data.GURU.NAMA_LENGKAP,
                'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
            }, guru)

            if not (exist):
                guru.append({
                    'GURU': data.GURU.NAMA_LENGKAP,
                    'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
                })

        jumlah = {}

        for data in kelas:
            jumlah[str(data)] = ''
        
        jumlah['JUMLAH_JAM_ATAU_RASIO'] = 0
        jumlah['JAM_TAMBAHAN'] = ''
        jumlah['TOTAL_JAM'] = 0
        jumlah['TUGAS_TAMBAHAN'] = ''
        
        # pembagian jadwal mengajar
        jumlah_keseluruhan = []
        for g in guru:
            if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
                data_baru = g
                data_baru = gabung_dictionary(data_baru, jumlah)
            else:
                data_baru = next(d for d in jumlah_keseluruhan if d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'])

            for data in jadwal:
                if data.GURU.NAMA_LENGKAP == g['GURU'] and str(data.MATA_PELAJARAN) == g['MATA_PELAJARAN']:
                    data_baru['JUMLAH_JAM_ATAU_RASIO'] += data.JUMLAH
    
                    if data_baru[str(data.KELAS)] == '':
                        data_baru[str(data.KELAS)] = data.JUMLAH
                    else:
                        data_baru[str(data.KELAS)] += data.JUMLAH
                        
                if data.GURU.NAMA_LENGKAP == g['GURU']:
                    data_baru['TOTAL_JAM'] += data.JUMLAH

            if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
                jumlah_keseluruhan.append(data_baru)
    
    
        # total kelas
        jumlah_siswa = {}
        for data in kelas:
            jumlah = kelas_siswa.filter(KELAS = data).count()    
            jumlah_siswa[str(data)] = jumlah
            
        
        for data in kelas:
            jumlah_siswa[str(data.KELAS.TINGKATAN)] = kelas_siswa.filter(KELAS__KELAS__TINGKATAN = data.KELAS.TINGKATAN).count()
        
        jumlah_total = []
        jumlah_total.append(jumlah_siswa)
        
        
        df_ = pd.DataFrame(jumlah_keseluruhan)
        df = pd.DataFrame(jumlah_keseluruhan).set_index(['GURU', 'MATA_PELAJARAN'])
        dt = pd.DataFrame(jumlah_total)
        
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine='openpyxl')
        
        df.to_excel(writer, sheet_name=self.KATEGORI, index=True, )
        dt.to_excel(writer, sheet_name='TOTAL_JUMLAH', index=False,)
        
        worksheet = writer.sheets[self.KATEGORI]
        worksheet2 = writer.sheets['TOTAL_JUMLAH']
        
        # col = get_columns_from_worksheet(worksheet)
        # total_jam = col['TOTAL_JAM']['number'] + 1
        
        # for car in df_['GURU'].unique():
        #     # find indices and add one to account for header
        #     u=df_.loc[df_['GURU']==car].index.values + 1

        #     if len(u) < 2:
        #         pass # do not merge cells if there is only one car name
        #     else:
        #         worksheet.merge_cells(start_row=u[0] + 1, start_column=total_jam, end_row=u[-1] + 1, end_column=total_jam)
        
        merge_cell(worksheet, df_, 'GURU', 'TOTAL_JAM')
                
    #     wb_obj = openpyxl.load_workbook(self.TEMPLATE.path)
    #     sheet_obj = wb_obj.active
        
    #     start_row = self.START_ROW - 2
    #     start_col = self.START_COL - 1
        
    #     kelas_row = start_row - 2
    #     kelas_col = start_col + 2
        
    #     mr = worksheet.max_row + 1
    #     mc = worksheet.max_column + 1
        
    #     mr2 = worksheet2.max_row + 1
    #     mc2 = worksheet2.max_column + 1
        
    #     no = 1
        
    #     for i in range (2, mr ):
    #         for j in range (1, mc ):
    #             # reading cell value from source excel file
    #             c = worksheet.cell(row = i, column = j)

    #             # writing the read value to destination excel file
    #             cell = sheet_obj.cell(row = i + start_row, column = start_col + j)
    #             cell.value = c.value
                
    #             if c.value != None and j == 1:
    #                 cell = sheet_obj.cell(row = i + kelas_row + 2, column = 1)
    #                 cell.value = no
    #                 no += 1
                
    #     for i in range (2, mr2 ):
    #         for j in range (1, mc2 ):
    #             # reading cell value from source excel file
    #             c = worksheet2.cell(row = i, column = j)

    #             # writing the read value to destination excel file
    #             cell = sheet_obj.cell(row = i + kelas_row, column = kelas_col + j)
    #             cell.value = c.value
                    
    #     virtual_workbook = BytesIO()
    #     wb_obj.save(virtual_workbook)
    #     self.FILE_HASIL = ContentFile(virtual_workbook.getvalue(), self.KATEGORI + '.xlsx')
        
        writer.save()
    
    elif self.KATEGORI == 'Rekapitulasi Jumlah Jam Mengajar':
        kelas_tingkatan = Kelas.objects.all().order_by('TINGKATAN',)
        jadwal = JadwalMengajar.objects.all().filter(TAHUN_AJARAN=self.TAHUN_AJARAN, SEMESTER = self.SEMESTER).order_by('GURU').annotate(JUMLAH=Count('WAKTU_PELAJARAN'))
        kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
        
        guru = []
        for data in jadwal:
            exist = multiple_value_exist({
                'GURU': data.GURU.NAMA_LENGKAP,
                'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
            }, guru)
            
            if not (exist):
                guru.append({
                    'GURU': data.GURU.NAMA_LENGKAP,
                    'NIP': data.GURU.NIP,
                    'PANGKAT': data.GURU.PANGKAT,
                    'GOL': cari_pangkat(data.GURU),
                    'MATA_PELAJARAN': str(data.MATA_PELAJARAN)
                })
                
        jumlah = {}
        
        for data in kelas_tingkatan:
            jumlah[str(data.TINGKATAN)] = 0

        jumlah_keseluruhan = []
        for g in guru:
            if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
                data_baru = g
                data_baru = gabung_dictionary(data_baru, jumlah)
            else:
                data_baru = next(d for d in jumlah_keseluruhan if d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'])

            for data in jadwal:
                if data.GURU.NAMA_LENGKAP == g['GURU'] and str(data.MATA_PELAJARAN) == g['MATA_PELAJARAN']:
                    tingkatan = data.KELAS.KELAS.TINGKATAN
                    data_baru[tingkatan] += data.JUMLAH

            if not any(d['GURU'] == g['GURU'] and d['MATA_PELAJARAN'] == g['MATA_PELAJARAN'] for d in jumlah_keseluruhan):
                jumlah_keseluruhan.append(data_baru)
    
        df = pd.DataFrame(jumlah_keseluruhan)
        df_ = pd.DataFrame(jumlah_keseluruhan).set_index(['GURU', 'NIP',])
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine='openpyxl')
        df_.to_excel(writer, sheet_name=self.KATEGORI, index=True, )
        print(df_)
        
        worksheet = writer.sheets[self.KATEGORI]
        merge_cell(worksheet, df, 'GURU', 'NIP')
        merge_cell(worksheet, df, 'GURU', 'PANGKAT')
        merge_cell(worksheet, df, 'GURU', 'GOL')
        
        # wb_obj = openpyxl.load_workbook(self.TEMPLATE.path)
        # sheet_obj = wb_obj.active
        
        # start_row = self.START_ROW - 2
        # start_col = self.START_COL - 1
        
        # no_row = start_row - 2
        # no_col = start_col + 2
        
        # mr = worksheet.max_row + 1
        # mc = worksheet.max_column + 1
        
        # no = 1
        
        # for i in range (2, mr ):
        #     for j in range (1, mc ):
        #         # reading cell value from source excel file
        #         c = worksheet.cell(row = i, column = j)

        #         # writing the read value to destination excel file
        #         cell = sheet_obj.cell(row = i + start_row, column = start_col + j)
        #         cell.value = c.value
                
        #         if c.value != None and j == 1:
        #             cell = sheet_obj.cell(row = i + no_row + 2, column = 1)
        #             cell.value = no
        #             no += 1
        
        # virtual_workbook = BytesIO()
        # wb_obj.save(virtual_workbook)
        # self.FILE_HASIL = ContentFile(virtual_workbook.getvalue(), self.KATEGORI + '.xlsx')
        
        
        writer.save()

#Pembagian Tugas BK

    elif self.KATEGORI == 'Pembagian Tugas BK Semester':
        pembagian_tugas = OfferingKelas.objects.all().order_by('KELAS__TINGKATAN',)
        
        kelas_tingkatan = OfferingKelas.objects.all().order_by('KELAS__TINGKATAN',)
        kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
        
        guru = []
        for data in pembagian_tugas:
            exist = multiple_value_exist({
                'NAMA_NIP' : data.DATA_GURU.NAMA_LENGKAP
            }, guru)
            
            if not (exist):
                guru.append({
                    'NAMA_NIP' : data.DATA_GURU.NAMA_LENGKAP
                })
        
        jumlah_keseluruhan = []
        
        for g in guru :
            if not any(d['NAMA_NIP'] == g['NAMA_NIP'] for d in jumlah_keseluruhan):
                jumlah_keseluruhan.append(g)
                
        
        kb = pd.DataFrame(jumlah_keseluruhan)
        print (kb)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine='openpyxl')
        kb.to_excel(writer, index=False)
        # worksheet = writer.sheets['BK']
        # print('worksheet')
        # print(df)
        writer.save()
        
        
        # guru = []
        # for data in pembagian_tugas:
        #     exist = multiple_value_exist({
        #         'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP + ',' + data.DATA_GURU.NIP
        #     }, guru)
            
        #     if not (exist):
        #         guru.append({
        #             'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP + ',' + data.DATA_GURU.NIP,
        #             'PANGKAT_GOLONGAN': data.DATA_GURU.PANGKAT + ',' + str(cari_pangkat(data.DATA_GURU)),
        #         })
        
                
        # jumlah = {}
        # jumlah_keseluruhan = []
        
        # for g in guru:
        #     if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
        #         data_baru = g
        #         data_baru = gabung_dictionary(data_baru, jumlah)
        #     else:
        #         data_baru = next(d for d in jumlah_keseluruhan if d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'])

        #     list_kelas_x =  []
        #     list_kelas_xi =  []
        #     list_kelas_xii =  []
            
        #     jumlah_kelas_x = []
        #     jumlah_kelas_xi = []
        #     jumlah_kelas_xii = []
        #     for data in pembagian_tugas:
        #         for data_kelas in data.DATA_KELAS.all():
        #             if data_kelas.KELAS.TINGKATAN == 'X' :
        #                 list_kelas_x.append(str(data_kelas))
        #                 jumlah_kelas_x.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
        #             elif data_kelas.KELAS.TINGKATAN == 'XI' :
        #                 list_kelas_xi.append(str(data_kelas))               
        #                 jumlah_kelas_xi.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
        #             elif data_kelas.KELAS.TINGKATAN == 'XII' :
        #                 list_kelas_xii.append(str(data_kelas))               
        #                 jumlah_kelas_xii.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
        #             # jumlah[str(data)]  = kelas_siswa.filter(KELAS = data_kelas,).count()

        #     data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        #     data_baru['JUMLAH_KELAS_X'] = ", ".join(jumlah_kelas_x)
        #     data_baru['KELAS_XI'] = ", ".join(list_kelas_xi)
        #     data_baru['JUMLAH_KELAS_XI'] = ", ".join(jumlah_kelas_xi)
        #     data_baru['KELAS_XII'] = ", ".join(list_kelas_xii)
        #     data_baru['JUMLAH_KELAS_XII'] = ", ".join(jumlah_kelas_xii)
            
        #     data_baru['KETERANGAN'] = data.KETERANGAN_TUGAS
        #     if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
        #         jumlah_keseluruhan.append(data_baru)
        #         print('merdeka')
        
        # print (jumlah_keseluruhan)
        
        # df = pd.DataFrame(jumlah_keseluruhan)
        # b = BytesIO()
        # writer = pd.ExcelWriter(b, engine='openpyxl', )
        # df.to_excel(writer, sheet_name=self.KATEGORI , merge_cells=True)
        # worksheet = writer.sheets[self.KATEGORI]
        # print('worksheet')
        # print(df)
        # # writer.save()

    elif self.KATEGORI == 'Pembagian Tugas Pokok dan Tambahan Tenaga Kependidikan':
        pembagian_tugas_tendik = OfferingKelas.objects.all().order_by('KELAS__TINGKATAN',)
        
        
        print ('Wew')
        guru = []
        
        for data in pembagian_tugas_tendik:
            exist = multiple_value_exist({
                'NAMA': data.DATA_GURU.NAMA_LENGKAP,
            }, guru)
            print ('guru ; ' + str(data.DATA_GURU.NAMA_LENGKAP))
            
        
            if not (exist):
                guru.append({
                    'NAMA': data.DATA_GURU.NAMA_LENGKAP,
                    'NIP': data.DATA_GURU.NIP,
                    'GOL' : cari_pangkat(data),
                    'PANGKAT' : data.DATA_GURU.PANGKAT,
                    'TUGAS_POKOK' : data.TUGAS_POKOK.JENIS_TUGAS,
                    'TUGAS_TAMBAHAN' : data.TUGAS_TAMBAHAN,
                })
        
        jumlah = {}
        jumlah_keseluruhan = []
        
        for g in guru:
            if not any(d['NAMA'] == g['NAMA'] for d in jumlah_keseluruhan):
                data_baru = g
                data_baru = gabung_dictionary(data_baru)
            else:
                data_baru = next(d for d in jumlah_keseluruhan if d['NAMA'] == g['NAMA'])
                    
            if not any(d['NAMA'] == g['NAMA']  for d in jumlah_keseluruhan):
                jumlah_keseluruhan.append(data_baru)
                
        
                
        df = pd.DataFrame(guru)
        b = BytesIO()
        writer = pd.ExcelWriter(b, engine='openpyxl')
        df.to_excel(writer, sheet_name='TUGAS_TAMBAHAN', index=False, )
        print('df :' + str(df))
        print ('Tes Darah')
        writer.save()
        print ('Tes Merdeka')
        self.FILE_HASIL = ContentFile(b.getvalue(), 'TEST' + '.xlsx')
                        
        
    # # check execution time -- finish
    # print('Waktu Eksekusi: ' + str(time.time() - start))
    
    return ContentFile(b.getvalue(), self.KATEGORI + '.xlsx')

            
    

def pembagian_bk(self, count):
    pembagian_tugas = count 
    kelas_siswa = KelasSiswa.objects.all().filter(KELAS__KELAS__TAHUN_AJARAN = self.TAHUN_AJARAN)
    # for tets in pembagian_tugas:
    #     print (tets.DATA_KELAS.all)
    # bk = []
    # for data in pembagian_tugas:
    #     list_kelas = {} 
    #     for kelas in data.DATA_KELAS.all(): 
    #         if kelas.KELAS.TINGKATAN == 'X': 
    #             list_kelas['X'] = kelas
    #         elif kelas.KELAS.TINGKATAN == 'XI': 
    #             list_kelas['XI'] = kelas
    #         elif kelas.KELAS.TINGKATAN == 'XII': 
    #             list_kelas['XII'] = kelas
    #     print (list_kelas)
    data_pembagian =[]
    data_cek = {}
    guru = []
    for data in pembagian_tugas:
        exist = multiple_value_exist({
            'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
        }, guru)
        
        if not (exist):
            guru.append({
                'NAMA_NIP_JABATAN': data.DATA_GURU.NAMA_LENGKAP,
                'PANGKAT_GOLONGAN': data.DATA_GURU.PANGKAT + ',' + str(cari_pangkat(data.DATA_GURU)),
            })
    
            
    jumlah = {}
    jumlah_keseluruhan = []
    
    for g in guru:
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            data_baru = g
            data_baru = gabung_dictionary(data_baru, jumlah)
        else:
            data_baru = next(d for d in jumlah_keseluruhan if d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'])
            
        
        list_kelas_x =  []
        list_kelas_xi =  []
        list_kelas_xii =  []
        
        jumlah_kelas_x = []
        jumlah_kelas_xi = []
        jumlah_kelas_xii = []
        # tugas_individual = pembagian_tugas.filter(DAT)
        for data in pembagian_tugas.filter(DATA_GURU__NAMA_LENGKAP = g['NAMA_NIP_JABATAN']):
            for data_kelas in data.DATA_KELAS.all():
                print (data_kelas)
                if data_kelas.KELAS.TINGKATAN == 'X' :
                    print (data_kelas)
                    list_kelas_x.append(str(data_kelas))
                    jumlah_kelas_x.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XI' :
                    list_kelas_xi.append(str(data_kelas))               
                    jumlah_kelas_xi.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                elif data_kelas.KELAS.TINGKATAN == 'XII' :
                    list_kelas_xii.append(str(data_kelas))               
                    jumlah_kelas_xii.append(str(kelas_siswa.filter(KELAS = data_kelas).count()))
                # jumlah[str(data)]  = kelas_siswa.filter(KELAS = data_kelas,).count()

        # data_baru['KELAS_X'] = ", ".join(list_kelas_x)
        data_baru['KELAS_X'] = "~ ".join(list_kelas_x)
        data_baru['JUMLAH_KELAS_X'] = "~ ".join(jumlah_kelas_x)
        data_baru['KELAS_XI'] = "~ ".join(list_kelas_xi)
        data_baru['JUMLAH_KELAS_XI'] = "~ ".join(jumlah_kelas_xi)
        data_baru['KELAS_XII'] = "~ ".join(list_kelas_xii)
        data_baru['JUMLAH_KELAS_XII'] = "~ ".join(jumlah_kelas_xii)
        
        merdeka = data_baru['KELAS_X'].split('~')
        # print ('test' + str(merdeka[5]))
        
        # for list in data_baru:
        data_x = data_baru['KELAS_X'].split('~')
        data_xi = data_baru['KELAS_XI'].split('~')
        data_xii = data_baru['KELAS_XII'].split('~')
        # print ('data_x_test' + str(data_x[0]))
        # print ('data_x_test' + str(data_x[1]))
        # print ('data_x_test' + str(data_x[0]))
        for i in range(5):
            leng_x = ((int(len(data_x)))-i)
            leng_xi = ((int(len(data_xi)))-i)
            leng_xii = ((int(len(data_xii)))-i)
            print ((int(len(data_x))))
            data_cek['NAMA_NIP_JABATAN'] = g['NAMA_NIP_JABATAN']
            if leng_x > 0 and leng_x <= (int(len(data_x))): 
                data_cek['KELAS_X'] = data_x[leng_x-1]
                print ('Kelas_yang_dicari' + str(data_x[leng_x-1]))
            else :
                data_cek['KELAS_X'] = ' '
            if leng_xi > 0: 
                data_cek['KELAS_XI'] = data_xi[leng_xi-1]
            else :
                data_cek['KELAS_XI'] = ' '
            if leng_xii > 0: 
                data_cek['KELAS_XII'] = data_xii[leng_xii - 1]
            else :
                data_cek['KELAS_XII'] = ' '
            data_pembagian.append(data_cek)
        print ('a')
        print(data_pembagian)
        print('b')
        data_baru['KETERANGAN'] = data.KETERANGAN_TUGAS
        if not any(d['NAMA_NIP_JABATAN'] == g['NAMA_NIP_JABATAN'] for d in jumlah_keseluruhan):
            jumlah_keseluruhan.append(data_baru)
    
            # print('merdeka')
    # hasil = []
    # tets_2 = []
    # for next_data in jumlah_keseluruhan :
    #     kelas_x = next_data['KELAS_X'].split(',')
    #     print (kelas_x[0])
    #     for i in range(len(kelas_x)) : 
    #         index = i - 1
    #         data_kelas = (kelas_x[i])
    #         tets_2['KELAS_X'] = data_kelas
            
    # print (hasil)
    # print (str(jumlah_kelas_x[0].split(',')))
    print (jumlah_keseluruhan)
    dt = pd.DataFrame(jumlah_keseluruhan)
    # dt = dt.applymap(str)
    # dt = dt.astype(str)
    b = BytesIO()
    writer = pd.ExcelWriter(b, engine='openpyxl')
    dt.to_excel(writer, sheet_name='PEMBAGIAN_BK', index=True)
    # dt.to_excel("output.xlsx")  
    writer.save()
    worksheet = writer.sheets['PEMBAGIAN_BK']
    # wb = Workbook(dt)
    # wb.save('balances.xlsx')
    wb_obj = openpyxl.load_workbook(self.TEMPLATE.path)
    sheet_obj = wb_obj.active
        
    start_row = self.START_ROW - 2
    start_col = self.START_COL - 1
    
    no_row = start_row - 2
    no_col = start_col + 2
    
    mr = worksheet.max_row + 1
    mc = worksheet.max_column + 1
    
    no = 1
    
    for i in range (2, mr ):
        baris_terbanyak = 0
        
        for j in range (1, mc ):
            # reading cell value from source excel file
            c = worksheet.cell(row = i, column = j)
            s = str(c.value).split('~')
            if len(s) > baris_terbanyak:
                baris_terbanyak = len(s)
                print (str(i) + '-' + str(j))
                
                
            # writing the read value to destination excel file
            cell = sheet_obj.cell(row = i + start_row, column = start_col + j)
            cell.value = c.value
            
            cell = sheet_obj.cell(row = i + start_row, column = start_col + 3)
            cell.value = s[0]
            
            
            # if c.value != None and j == 1:
            #     cell = sheet_obj.cell(row = i + no_row + 2, column = 1)
            #     cell.value = no
            #     no += 1
        print (i, j)
        print(baris_terbanyak)
    
    
    
    virtual_workbook = BytesIO()
    wb_obj.save(virtual_workbook)
    self.FILE_HASIL = ContentFile(virtual_workbook.getvalue(), self.KATEGORI + '.xlsx')
    
        
    return ContentFile(b.getvalue(), 'PEMBAGIAN_BK' + '.xlsx') 

